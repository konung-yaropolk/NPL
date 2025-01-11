import sys
import matplotlib.pyplot as plt
import numpy as np
import csv
import statlib
from openpyxl import load_workbook
from PyQt6.QtWidgets import QApplication, QMainWindow, QPlainTextEdit
# , QMessageBox, QTextEdit, QPushButton
from mainwindow_layout import Ui_MainWindow


class miscMethods():

    def runAuto(self):
        self.ui.errorMsg.setText('')
        data_text = self.ui.input_field.toPlainText()
        try:
            groups_list = self.listify_text(data_text)
            analysis = statlib.StatisticalAnalysis(
                groups_list,
                paired=self.ui.dependendGroups.isChecked(),
                tails=2 if self.ui.twoTailed.isChecked() else 1,
                popmean=float(self.ui.popMean.text()))

            analysis.RunAuto()
            results = analysis.GetResult()
            summary = analysis.GetSummary()

            self.ui.result_display.setPlainText(summary)
            self.initiate_barplot(results)
        except Exception as e:
            e = 'Error: \n' + str(e)
            print(e)
            self.ui.errorMsg.setText(e)

    def runManual(self):
        self.ui.errorMsg.setText('')
        data_text = self.ui.input_field.toPlainText()
        try:
            groups_list = self.listify_text(data_text)
            analysis = statlib.StatisticalAnalysis(
                groups_list,
                paired=self.ui.dependendGroups.isChecked(),
                tails=2 if self.ui.twoTailed.isChecked() else 1,
                popmean=float(self.ui.popMean.text()))

            checkbuttons_state_list = [
                self.ui.anova.isChecked(),
                self.ui.friedman.isChecked(),
                self.ui.kruskal_wallis.isChecked(),
                self.ui.mann_whitney.isChecked(),
                self.ui.t_test_independend.isChecked(),
                self.ui.t_test_paired.isChecked(),
                self.ui.t_test_single_sample.isChecked(),
                self.ui.wilcoxon.isChecked(),
                self.ui.wilcoxon_single_sample.isChecked(),
            ]

            checked_test_ids = []
            for i in range(len(analysis.test_ids_all)):
                if checkbuttons_state_list[i]:
                    checked_test_ids.append(analysis.test_ids_all[i])

            for test_id in checked_test_ids:
                analysis.RunManual(test_id)
                results = analysis.GetResult()
                self.initiate_barplot(results)

            summary = analysis.GetSummary()
            self.ui.result_display.setPlainText(summary)

        except Exception as e:
            e = 'Error: \n' + str(e)
            print(e)
            self.ui.errorMsg.setText(e)

    def initiate_barplot(self, results):
        if results and self.ui.makePlot.isChecked():
            self.barplot(results['Samples'],
                         p=results['p-value'],
                         stars=results['Stars_Printed'],
                         sd=results['Groups_SD'],
                         mean=results['Groups_Mean'],
                         median=results['Groups_Median'],
                         testname=results['Test_Name'],
                         n=results['Groups_N'],
                         )

    def barplot(self, data_samples, p=1, stars='ns', sd=0, mean=0, median=0, testname='', n=0):
        fig, ax = plt.subplots(figsize=(3, 4))

        colors = ['k', 'r', 'b', 'g']
        colors_fill = ['#CCCCCC', '#FFCCCC', '#CCCCFF', '#CCFFCC']

        for i, data in enumerate(data_samples):
            x = i + 1  # Bar position
            # Bars:
            ax.bar(x,
                   mean[i],
                   yerr=sd[i],
                   width=.8,
                   capsize=10,
                   ecolor='r',
                   edgecolor=colors[i % len(colors)],
                   facecolor=colors_fill[i % len(colors_fill)],
                   fill=True,
                   linewidth=2)
            # Data points:
            # Adjust random horizontal spread range
            spread = np.random.uniform(-.10, .10, size=len(data))
            ax.scatter(x + spread, data, color='black',
                       s=16, zorder=1, alpha=0.5)
            ax.plot(x,
                    median[i],
                    marker='x',
                    markerfacecolor='#00000000',
                    markeredgecolor='r',
                    markersize=10,
                    markeredgewidth=1)

        # Significance bar and stars
        y_range = max([max(data) for data in data_samples])
        x1, x2 = 1, len(data_samples)
        y, h, col = 1.05 * y_range, .05 * y_range, 'k'
        ax.plot([x1, x1, x2, x2], [y, y + h, y + h, y], lw=1.5, c=col)
        ax.text((x1 + x2) * .5,
                y + h,
                '{}\n{}'.format(p, stars),
                ha='center',
                va='bottom',
                color=col)

        # Add subtitle
        fig.text(0.95, 0.01, '{}\nn={}'.format(testname, str(n)[1:-1]),
                 ha='right', va='bottom', fontsize=8)

        # Remove borders
        for spine in ax.spines.values():
            spine.set_visible(False)
        ax.spines['left'].set_visible(True)
        ax.xaxis.set_visible(False)

        plt.show()

    def listify_text(self, data_text):
        matrix = [list(group.split('\t')) for group in data_text.split('\n')]
        max_len = max(len(row) for row in matrix)
        padded_matrix = [row + [None] * (max_len - len(row)) for row in matrix]
        # transpose the matrix:
        transposed = [[padded_matrix[j][i]
                       for j in range(len(padded_matrix))] for i in range(max_len)]
        # Remove None values if padding was used
        return [[element for element in row if element is not None] for row in transposed]


class PlainTextEditDragNDrop(QPlainTextEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.endswith('.csv'):
                self.open_csv(file_path)
            elif file_path.endswith('.xlsx'):
                self.open_xlsx(file_path)
            else:
                self.setPlainText("Only CSV and XLSX files are supported.")

    def open_csv(self, file_path):
        with open(file_path, 'r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            content = '\n'.join(['\t'.join(row) for row in reader])
            self.setPlainText(content)

    def open_xlsx(self, file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        content = '\n'.join(['\t'.join(
            [str(cell.value) if cell.value is not None else '' for cell in row]) for row in sheet.iter_rows()])
        self.setPlainText(content)


class MainWindow(QMainWindow, miscMethods):

    def __init__(self):
        super().__init__()
        self.load_UI()

    def load_UI(self,):
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        # Bind functions to butts:
        self.ui.runAutoButton.clicked.connect(self.runAuto)
        self.ui.runManualButton.clicked.connect(self.runManual)
        # replace the input_field with Drag-n-Drop-able one:
        self.replace_input_field()

    def replace_input_field(self):
        new_text_edit = PlainTextEditDragNDrop(self)
        placeholder_text = self.ui.input_field.placeholderText()
        layout = self.centralWidget().layout()
        layout.replaceWidget(self.ui.input_field, new_text_edit)
        self.ui.input_field.deleteLater()
        self.ui.input_field = new_text_edit
        self.ui.input_field.setPlaceholderText(placeholder_text)


if __name__ == '__main__':
    # import pyi_splash
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    # pyi_splash.close()
    sys.exit(app.exec())
